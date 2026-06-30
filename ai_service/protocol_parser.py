from __future__ import annotations

import re
import csv
import json
from io import BytesIO
from pathlib import Path

import pdfplumber
from docx import Document
from docx.document import Document as DocxDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook


SUPPORTED_PROTOCOL_EXTENSIONS = {".pdf", ".docx", ".txt", ".json", ".csv", ".xlsx"}


def extract_protocol_text(filename: str, content: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_PROTOCOL_EXTENSIONS:
        raise ValueError("Unsupported file type. Use PDF, DOCX, or TXT.")

    if ext == ".pdf":
        return _extract_pdf_text(content)
    if ext == ".docx":
        return _extract_docx_text(content)
    if ext == ".json":
        return _extract_json_text(content)
    if ext == ".csv":
        return _extract_csv_text(content)
    if ext == ".xlsx":
        return _extract_xlsx_text(content)
    return _extract_txt_text(content)


def _extract_pdf_text(content: bytes) -> str:
    chunks: list[str] = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text.strip():
                chunks.append(page_text.strip())

    text = _normalize_protocol_text("\n\n".join(chunks))
    if not text:
        raise ValueError("No readable text found in PDF.")
    return text


def _extract_docx_text(content: bytes) -> str:
    doc = Document(BytesIO(content))
    chunks: list[str] = []

    for block in _iter_docx_blocks(doc):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if text and (not chunks or chunks[-1] != text):
                chunks.append(text)
            continue

        if isinstance(block, Table):
            for row in block.rows:
                cells = [_extract_cell_text(cell) for cell in row.cells]
                cells = [cell for cell in cells if cell]
                if not cells:
                    continue
                line = " | ".join(cells)
                if not chunks or chunks[-1] != line:
                    chunks.append(line)

    text = _normalize_protocol_text("\n\n".join(chunks))
    if not text:
        raise ValueError("No readable text found in DOCX.")
    return text


def _extract_txt_text(content: bytes) -> str:
    text = _normalize_protocol_text(content.decode("utf-8", errors="replace"))
    if not text:
        raise ValueError("Uploaded TXT file is empty.")
    return text


def _extract_json_text(content: bytes) -> str:
    raw = content.decode("utf-8", errors="replace")
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as error:
        raise ValueError(f"Uploaded JSON is invalid: {error.msg}")

    text = _normalize_protocol_text(json.dumps(parsed, indent=2, ensure_ascii=False))
    if not text:
        raise ValueError("Uploaded JSON file is empty.")
    return text


def _extract_csv_text(content: bytes) -> str:
    raw = content.decode("utf-8-sig", errors="replace")
    rows = []
    for row in csv.reader(raw.splitlines()):
        cells = [cell.strip() for cell in row if cell.strip()]
        if cells:
            rows.append(" | ".join(cells))

    text = _normalize_protocol_text("\n".join(rows))
    if not text:
        raise ValueError("Uploaded CSV file is empty.")
    return text


def _extract_xlsx_text(content: bytes) -> str:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    chunks: list[str] = []

    for sheet in workbook.worksheets:
        chunks.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell not in (None, "")]
            if cells:
                chunks.append(" | ".join(cells))

    text = _normalize_protocol_text("\n".join(chunks))
    if not text:
        raise ValueError("Uploaded XLSX file is empty.")
    return text


def _iter_docx_blocks(doc: DocxDocument):
    for child in doc.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, doc)
        elif isinstance(child, CT_Tbl):
            yield Table(child, doc)


def _extract_cell_text(cell) -> str:
    parts = [paragraph.text.strip() for paragraph in cell.paragraphs if paragraph.text.strip()]
    return " ".join(parts).strip()


def _normalize_protocol_text(text: str) -> str:
    cleaned = text.replace("\u25cf", "-")
    cleaned = cleaned.replace("\u2022", "-")
    cleaned = cleaned.replace("\u2013", "-").replace("\u2014", "-")
    cleaned = cleaned.replace("\r\n", "\n").replace("\r", "\n")
    cleaned = re.sub(r"-\s*-\s*", "- ", cleaned)
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()
